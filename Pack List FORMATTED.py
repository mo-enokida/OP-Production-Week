import openpyxl
wb = openpyxl.load_workbook('6. Shipping Report Packlist OUTPUT.xlsx')
from openpyxl.styles import Font, Color, Fill, NamedStyle, PatternFill, Border, Side, Alignment, Protection
from openpyxl.styles import Font, NamedStyle, PatternFill, Border, Side, Alignment, Protection
from openpyxl.cell import Cell
from copy import copy, deepcopy



###DONT COPY
red_meal = 'Italian Braised Chicken'
teal_meal = 'Chicken Lo Mein'
green_meal = 'Sweet Corn Ravioli'
yellow_meal = 'Potato Spinach Dosas'
blue_meal = 'Vegetable Lo Mein'


####DONT COPY

##TOTAL NUMBERS

sheet = wb['TOTAL_NUMBERS']

for row in range(2, sheet.max_row + 1):
    gf_cell = sheet.cell(row, 3)
    if gf_cell.value == "X" or gf_cell.value == "x":
        gf_cell.value = "Gluten Free"

meal_column = 6
for row in range(2, sheet.max_row + 1):
    meal_cell = sheet.cell(row, meal_column)
    meals = str(meal_cell.value)
    if blue_meal or yellow_meal or green_meal or teal_meal or red_meal in meals:
        meals = meals.replace(blue_meal, 'blue')
        meals = meals.replace(yellow_meal, 'yellow')
        meals = meals.replace(green_meal, 'green')
        meals = meals.replace(teal_meal, 'teal')
        meals = meals.replace(red_meal, 'red')
        meal_cell.value = meals
    else:
        meal_column = 6



###REG SMOOTHIES
reg_sheet = wb['ShipRep_REG']
list_sheet = wb['REG_ADDONS']
addon_list = ['S1', 'S2', 'S3', 'S4', 'S5', 'S6', 'S7', 'S8', 'S9', 'S10','S11','S12','S13','S14','S15','S16','S17','S18','S19','S20','S21','S22','S23','S24','S25',
              'S26','S27','S28','S29','S30','S31','S32','S33','S34','S35','S36','S37','S38','S39','S40','S41','S42','S43','S44','S45',
              'S46','S47','S48','S49','S50','S51','S52','S53','S54','S55','S56','S57','S58','S59','S60','S61','S62','S63','S64','S65',
              'S66','S67','S68','S69','S70','S71','S72','S73','S74','S75','S76','S77','S78','S79','S80','S81','S82','S83','S84','S85',
              'S86','S87','S88','S89','S90','S91','S92','S93','S94','S95','S96','S97','S98','S99','S100','S101','S102','S103','S104','S105',
              'S106','S107','S108','S109','S110','S111','S112','S113','S114','S115','S116','S117','S118','S119','S120','S121','S122','S123','S124','S125',
              'S126','S127','S128','S129','S130','S131','S132','S133','S134','S135','S136','S137','S138','S139','S140','S141','S142','S143','S144','S145',
              'S146','S147','S148','S149','S150','S151','S152','S153','S154','S155','S156','S157','S158','S159','S160','S161','S162','S163','S164','S165',
              'S166','S167','S168','S169','S170','S171','S172','S173','S174','S175','S176','S177','S178','S179','S180','S181','S182','S183','S184','S185',
              'S186','S187','S188','S189','S190','S191','S192','S193','S194','S195','S196','S197','S198','S199','S200','S201','S202','S203','S204','S205',
              'S206','S207','S208','S209']

alpha_column = 26
name_column = 1
kids_column = 4
packlist_row = 4
smoothie_column = 6
antiox_column = 8
bigred_column = 9
blue_column = 10
boost_column = 11
green_column = 12
pbj_column = 13
vitc_column = 14
cookies_column = 15
break_column = 17
chip_column = 18
chipgf_column = 19
or_column = 20
orgf_column = 21
sugar_column = 22
sugargf_column = 23

for row in range(2, reg_sheet.max_row +1):
    alpha_cell = reg_sheet.cell(row, alpha_column)
    name_cell = reg_sheet.cell(row, name_column)
    kids_cell = reg_sheet.cell(row,kids_column)
    smoothie_cell = reg_sheet.cell(row,smoothie_column)
    antiox_cell = reg_sheet.cell(row, antiox_column)
    bigred_cell = reg_sheet.cell(row, bigred_column)
    blue_cell = reg_sheet.cell(row, blue_column)
    boost_cell = reg_sheet.cell(row, boost_column)
    green_cell = reg_sheet.cell(row, green_column)
    pbj_cell = reg_sheet.cell(row, pbj_column)
    vitc_cell = reg_sheet.cell(row,vitc_column)
    cookies_cell = reg_sheet.cell(row, cookies_column)
    break_cell = reg_sheet.cell(row, break_column)
    chip_cell = reg_sheet.cell(row, chip_column)
    chipgf_cell = reg_sheet.cell(row, chipgf_column)
    or_cell = reg_sheet.cell(row, or_column)
    orgf_cell = reg_sheet.cell(row, orgf_column)
    sugar_cell = reg_sheet.cell(row,sugar_column)
    sugargf_cell = reg_sheet.cell(row,sugargf_column)
    alpha = alpha_cell.value
    name = name_cell.value
    kids = kids_cell.value
    if alpha_cell.value in addon_list:
        pack_alpha = list_sheet.cell(packlist_row, 1)
        pack_kids = list_sheet.cell(packlist_row, 2)
        pack_name = list_sheet.cell(packlist_row, 3)
        pack_smoothies = list_sheet.cell(packlist_row, 4)
        pack_antiox = list_sheet.cell(packlist_row, 5)
        pack_bigred = list_sheet.cell(packlist_row, 6)
        pack_blue = list_sheet.cell(packlist_row, 7)
        pack_boost = list_sheet.cell(packlist_row, 8)
        pack_green = list_sheet.cell(packlist_row, 9)
        pack_pbj = list_sheet.cell(packlist_row, 10)
        pack_vitc = list_sheet.cell(packlist_row, 11)
        pack_cookie = list_sheet.cell(packlist_row, 12)
        pack_break = list_sheet.cell(packlist_row, 13)
        pack_chip = list_sheet.cell(packlist_row, 14)
        pack_chipgf = list_sheet.cell(packlist_row, 15)
        pack_or = list_sheet.cell(packlist_row, 16)
        pack_orgf = list_sheet.cell(packlist_row, 17)
        pack_sugar = list_sheet.cell(packlist_row, 18)
        pack_sugargf = list_sheet.cell(packlist_row, 19)
        pack_alpha.value = alpha
        pack_name.value = name
        pack_kids.value = kids
        pack_smoothies.value = smoothie_cell.value
        pack_antiox.value = antiox_cell.value
        pack_bigred.value = bigred_cell.value
        pack_blue.value = blue_cell.value
        pack_boost.value = boost_cell.value
        pack_green.value = green_cell.value
        pack_pbj.value = pbj_cell.value
        pack_vitc.value = vitc_cell.value
        pack_cookie.value = cookies_cell.value
        pack_break.value = break_cell.value
        pack_chip.value = chip_cell.value
        pack_chipgf.value = chipgf_cell.value
        pack_or.value = or_cell.value
        pack_orgf.value = orgf_cell.value
        pack_sugar.value = sugar_cell.value
        pack_sugargf.value = sugargf_cell.value
        packlist_row = packlist_row + 1
wb.save('7. CHECK.xlsx')
wb = openpyxl.load_workbook('7. CHECK.xlsx')
##REG SMOOTHIE COLORS
sheet = wb['REG_ADDONS']
fourk_cell = sheet.cell(1, 21)
threek_cell = sheet.cell(2, 21)
twok_cell = sheet.cell(3, 21)
onek_cell = sheet.cell(4, 21)
zerok_cell = sheet.cell(5, 21)
if fourk_cell.has_style:
    if 'FOURk_style' not in wb.named_styles:
        FOURk_style = NamedStyle(name='FOURk_style')
        FOURk_style.font = copy(fourk_cell.font)
        FOURk_style.border = copy(fourk_cell.border)
        FOURk_style.fill = copy(fourk_cell.fill)
        FOURk_style.number_format = copy(fourk_cell.number_format)
        FOURk_style.protection = copy(fourk_cell.protection)
        FOURk_style.alignment = copy(fourk_cell.alignment)
        wb.add_named_style(FOURk_style)
if threek_cell.has_style:
    if 'THREEk_style' not in wb.named_styles:
        THREEk_style = NamedStyle(name='THREEk_style')
        THREEk_style.font = copy(threek_cell.font)
        THREEk_style.border = copy(threek_cell.border)
        THREEk_style.fill = copy(threek_cell.fill)
        THREEk_style.number_format = copy(threek_cell.number_format)
        THREEk_style.protection = copy(threek_cell.protection)
        THREEk_style.alignment = copy(threek_cell.alignment)
        wb.add_named_style(THREEk_style)
if twok_cell.has_style:
    if 'TWOk_style' not in wb.named_styles:
        TWOk_style = NamedStyle(name='TWOk_style')
        TWOk_style.font = copy(twok_cell.font)
        TWOk_style.border = copy(twok_cell.border)
        TWOk_style.fill = copy(twok_cell.fill)
        TWOk_style.number_format = copy(twok_cell.number_format)
        TWOk_style.protection = copy(twok_cell.protection)
        TWOk_style.alignment = copy(twok_cell.alignment)
        wb.add_named_style(TWOk_style)
if onek_cell.has_style:
    if 'ONEk_style' not in wb.named_styles:
        ONEk_style = NamedStyle(name='ONEk_style')
        ONEk_style.font = copy(onek_cell.font)
        ONEk_style.border = copy(onek_cell.border)
        ONEk_style.fill = copy(onek_cell.fill)
        ONEk_style.number_format = copy(onek_cell.number_format)
        ONEk_style.protection = copy(onek_cell.protection)
        ONEk_style.alignment = copy(onek_cell.alignment)
        wb.add_named_style(ONEk_style)
if zerok_cell.has_style:
    if 'ZEROk_style' not in wb.named_styles:
        ZEROk_style = NamedStyle(name='ZEROk_style')
        ZEROk_style.font = copy(zerok_cell.font)
        ZEROk_style.border = copy(zerok_cell.border)
        ZEROk_style.fill = copy(zerok_cell.fill)
        ZEROk_style.number_format = copy(zerok_cell.number_format)
        ZEROk_style.protection = copy(zerok_cell.protection)
        ZEROk_style.alignment = copy(zerok_cell.alignment)
        wb.add_named_style(ZEROk_style)

kids_column = 2
for row in range(4, sheet.max_row + 1):
    kids = sheet.cell(row, kids_column)
    if kids.value == '4':
        for col in range(1, 4):
            Cell = sheet.cell(row, col)
            Cell.style = 'FOURk_style'
    elif kids.value == '3':
        for col in range(1,4):
            Cell = sheet.cell(row, col)
            Cell.style = 'THREEk_style'
    elif kids.value == '2':
        for col in range(1,4):
            Cell = sheet.cell(row, col)
            Cell.style = 'TWOk_style'
    elif kids.value == '1':
        for col in range(1,4):
            Cell = sheet.cell(row, col)
            Cell.style = 'ONEk_style'
    elif kids.value == '0':
        for col in range(1,4):
            Cell = sheet.cell(row, col)
            Cell.style = 'ZEROk_style'

for row in range(4, sheet.max_row + 1):
    for col in range(5, 11):
        Cell = sheet.cell(row, col)
        if Cell.value == 0:
            Cell.value == ''
    for col in range (13, 19):
        Cell = sheet.cell(row, col)
        if Cell.value == 0:
            Cell.value == ''

wb.save('7. CHECK.xlsx')


###GF SMOOTHIES
reg_sheet = wb['ShipRep_GF']
list_sheet = wb['VIP.GF_ADDONS']
addon_list = ['GF1', 'GF2', 'GF3', 'GF4', 'GF5', 'GF6', 'GF7', 'GF8', 'GF9', 'GF10','GF11','GF12','GF13','GF14','GF15','GF16','GF17','GF18','GF19','GF20','GF21','GF22','GF23','GF24','GF25',
              'GF26','GF27','GF28','GF29','GF30','GF31','GF32','GF33','GF34','GF35','GF36','GF37','GF38','GF39','GF40','GF41','GF42','GF43','GF44','GF45',
              'GF46','GF47','GF48','GF49','GF50','GF51','GF52','GF53','GF54','GF55','GF56','GF57','GF58','GF59','GF60','GF61','GF62','GF63','GF64','GF65',
              'GF66','GF67','GF68','GF69','GF70','GF71','GF72','GF73','GF74','GF75','GF76','GF77','GF78','GF79','GF80','GF81','GF82','GF83','GF84','GF85',
              'GF86','GF87','GF88','GF89','GF90','GF91','GF92','GF93','GF94','GF95','GF96','GF97','GF98','GF99','GF100','GF101','GF102','GF103','GF104','GF105',
              'GF106','GF107','GF108','GF109','GF110','GF111','GF112','GF113','GF114','GF115','GF116','GF117','GF118','GF119','GF120','GF121','GF122','GF123','GF124','GF125',
              'GF126','GF127','GF128','GF129','GF130','GF131','GF132','GF133','GF134','GF135','GF136','GF137','GF138','GF139','GF140','GF141','GF142','GF143','GF144','GF145',
              'GF146','GF147','GF148','GF149','GF150','GF151','GF152','GF153','GF154','GF155','GF156','GF157','GF158','GF159','GF160','GF161','GF162','GF163','GF164','GF165',
              'GF166','GF167','GF168','GF169','GF170','GF171','GF172','GF173','GF174','GF175','GF176','GF177','GF178','GF179','GF180','GF181','GF182','GF183','GF184','GF185',
              'GF186','GF187','GF188','GF189','GF190','GF191','GF192','GF193','GF194','GF195','GF196','GF197','GF198','GF199','GF200','GF201','GF202','GF203','GF204','GF205',
              'GF206','GF207','GF208','GF209']

alpha_column = 26
name_column = 1
kids_column = 4
packlist_row = 22
smoothie_column = 6
antiox_column = 8
bigred_column = 9
blue_column = 10
boost_column = 11
green_column = 12
pbj_column = 13
vitc_column = 14
cookies_column = 15
break_column = 17
chip_column = 18
chipgf_column = 19
or_column = 20
orgf_column = 21
sugar_column = 22
sugargf_column = 23

for row in range(2, reg_sheet.max_row +1):
    alpha_cell = reg_sheet.cell(row, alpha_column)
    name_cell = reg_sheet.cell(row, name_column)
    kids_cell = reg_sheet.cell(row,kids_column)
    smoothie_cell = reg_sheet.cell(row,smoothie_column)
    antiox_cell = reg_sheet.cell(row, antiox_column)
    bigred_cell = reg_sheet.cell(row, bigred_column)
    blue_cell = reg_sheet.cell(row, blue_column)
    boost_cell = reg_sheet.cell(row, boost_column)
    green_cell = reg_sheet.cell(row, green_column)
    pbj_cell = reg_sheet.cell(row, pbj_column)
    vitc_cell = reg_sheet.cell(row,vitc_column)
    cookies_cell = reg_sheet.cell(row, cookies_column)
    break_cell = reg_sheet.cell(row, break_column)
    chip_cell = reg_sheet.cell(row, chip_column)
    chipgf_cell = reg_sheet.cell(row, chipgf_column)
    or_cell = reg_sheet.cell(row, or_column)
    orgf_cell = reg_sheet.cell(row, orgf_column)
    sugar_cell = reg_sheet.cell(row,sugar_column)
    sugargf_cell = reg_sheet.cell(row,sugargf_column)
    alpha = alpha_cell.value
    name = name_cell.value
    kids = kids_cell.value
    if alpha_cell.value in addon_list:
        pack_alpha = list_sheet.cell(packlist_row, 1)
        pack_kids = list_sheet.cell(packlist_row, 2)
        pack_name = list_sheet.cell(packlist_row, 3)
        pack_smoothies = list_sheet.cell(packlist_row, 4)
        pack_antiox = list_sheet.cell(packlist_row, 5)
        pack_bigred = list_sheet.cell(packlist_row, 6)
        pack_blue = list_sheet.cell(packlist_row, 7)
        pack_boost = list_sheet.cell(packlist_row, 8)
        pack_green = list_sheet.cell(packlist_row, 9)
        pack_pbj = list_sheet.cell(packlist_row, 10)
        pack_vitc = list_sheet.cell(packlist_row, 11)
        pack_cookie = list_sheet.cell(packlist_row, 12)
        pack_break = list_sheet.cell(packlist_row, 13)
        pack_chip = list_sheet.cell(packlist_row, 14)
        pack_chipgf = list_sheet.cell(packlist_row, 15)
        pack_or = list_sheet.cell(packlist_row, 16)
        pack_orgf = list_sheet.cell(packlist_row, 17)
        pack_sugar = list_sheet.cell(packlist_row, 18)
        pack_sugargf = list_sheet.cell(packlist_row, 19)
        pack_alpha.value = alpha
        pack_name.value = name
        pack_kids.value = kids
        pack_smoothies.value = smoothie_cell.value
        pack_antiox.value = antiox_cell.value
        pack_bigred.value = bigred_cell.value
        pack_blue.value = blue_cell.value
        pack_boost.value = boost_cell.value
        pack_green.value = green_cell.value
        pack_pbj.value = pbj_cell.value
        pack_vitc.value = vitc_cell.value
        pack_cookie.value = cookies_cell.value
        pack_break.value = break_cell.value
        pack_chip.value = chip_cell.value
        pack_chipgf.value = chipgf_cell.value
        pack_or.value = or_cell.value
        pack_orgf.value = orgf_cell.value
        pack_sugar.value = sugar_cell.value
        pack_sugargf.value = sugargf_cell.value
        packlist_row = packlist_row + 1
wb.save('7. CHECK.xlsx')
wb = openpyxl.load_workbook('7. CHECK.xlsx')
##GF SMOOTHIE COLORS
sheet = wb['VIP.GF_ADDONS']
fourk_cell = sheet.cell(1, 21)
threek_cell = sheet.cell(2, 21)
twok_cell = sheet.cell(3, 21)
onek_cell = sheet.cell(4, 21)
zerok_cell = sheet.cell(5, 21)
if fourk_cell.has_style:
    if 'FOURk_style' not in wb.named_styles:
        FOURk_style = NamedStyle(name='FOURk_style')
        FOURk_style.font = copy(fourk_cell.font)
        FOURk_style.border = copy(fourk_cell.border)
        FOURk_style.fill = copy(fourk_cell.fill)
        FOURk_style.number_format = copy(fourk_cell.number_format)
        FOURk_style.protection = copy(fourk_cell.protection)
        FOURk_style.alignment = copy(fourk_cell.alignment)
        wb.add_named_style(FOURk_style)
if threek_cell.has_style:
    if 'THREEk_style' not in wb.named_styles:
        THREEk_style = NamedStyle(name='THREEk_style')
        THREEk_style.font = copy(threek_cell.font)
        THREEk_style.border = copy(threek_cell.border)
        THREEk_style.fill = copy(threek_cell.fill)
        THREEk_style.number_format = copy(threek_cell.number_format)
        THREEk_style.protection = copy(threek_cell.protection)
        THREEk_style.alignment = copy(threek_cell.alignment)
        wb.add_named_style(THREEk_style)
if twok_cell.has_style:
    if 'TWOk_style' not in wb.named_styles:
        TWOk_style = NamedStyle(name='TWOk_style')
        TWOk_style.font = copy(twok_cell.font)
        TWOk_style.border = copy(twok_cell.border)
        TWOk_style.fill = copy(twok_cell.fill)
        TWOk_style.number_format = copy(twok_cell.number_format)
        TWOk_style.protection = copy(twok_cell.protection)
        TWOk_style.alignment = copy(twok_cell.alignment)
        wb.add_named_style(TWOk_style)
if onek_cell.has_style:
    if 'ONEk_style' not in wb.named_styles:
        ONEk_style = NamedStyle(name='ONEk_style')
        ONEk_style.font = copy(onek_cell.font)
        ONEk_style.border = copy(onek_cell.border)
        ONEk_style.fill = copy(onek_cell.fill)
        ONEk_style.number_format = copy(onek_cell.number_format)
        ONEk_style.protection = copy(onek_cell.protection)
        ONEk_style.alignment = copy(onek_cell.alignment)
        wb.add_named_style(ONEk_style)
if zerok_cell.has_style:
    if 'ZEROk_style' not in wb.named_styles:
        ZEROk_style = NamedStyle(name='ZEROk_style')
        ZEROk_style.font = copy(zerok_cell.font)
        ZEROk_style.border = copy(zerok_cell.border)
        ZEROk_style.fill = copy(zerok_cell.fill)
        ZEROk_style.number_format = copy(zerok_cell.number_format)
        ZEROk_style.protection = copy(zerok_cell.protection)
        ZEROk_style.alignment = copy(zerok_cell.alignment)
        wb.add_named_style(ZEROk_style)

kids_column = 2
for row in range(4, sheet.max_row + 1):
    kids = sheet.cell(row, kids_column)
    if kids.value == '4':
        for col in range(1, 4):
            Cell = sheet.cell(row, col)
            Cell.style = 'FOURk_style'
    elif kids.value == '3':
        for col in range(1,4):
            Cell = sheet.cell(row, col)
            Cell.style = 'THREEk_style'
    elif kids.value == '2':
        for col in range(1,4):
            Cell = sheet.cell(row, col)
            Cell.style = 'TWOk_style'
    elif kids.value == '1':
        for col in range(1,4):
            Cell = sheet.cell(row, col)
            Cell.style = 'ONEk_style'
    elif kids.value == '0':
        for col in range(1,4):
            Cell = sheet.cell(row, col)
            Cell.style = 'ZEROk_style'

for row in range(4, sheet.max_row + 1):
    for col in range(5, 11):
        Cell = sheet.cell(row, col)
        if Cell.value == 0:
            Cell.value == ''
    for col in range (13, 19):
        Cell = sheet.cell(row, col)
        if Cell.value == 0:
            Cell.value == ''

wb.save('7. CHECK.xlsx')


### ADD DELIVERY ROUTES BACK IN
wb = openpyxl.load_workbook('5. Shipping Report INPUT.xlsx')
fwb = openpyxl.load_workbook('7. Check.xlsx')
final_sheet = fwb['ShipRep_GF']
sheet = wb['ShipRep_GF']
delivery_column = 25
for row in range(2, sheet.max_row + 1):
    original_row = row
    delivery_cell = sheet.cell(row, delivery_column)
    overwrite_cell = final_sheet.cell(original_row, delivery_column)
    overwrite_cell.value = delivery_cell.value

fwb.save('7. CHECK.xlsx')

wb = openpyxl.load_workbook('5. Shipping Report INPUT.xlsx')
fwb = openpyxl.load_workbook('7. Check.xlsx')
final_sheet = fwb['ShipRep_REG']
sheet = wb['ShipRep_REG']
delivery_column = 25
for row in range(2, sheet.max_row + 1):
    original_row = row
    delivery_cell = sheet.cell(row, delivery_column)
    overwrite_cell = final_sheet.cell(original_row, delivery_column)
    overwrite_cell.value = delivery_cell.value

fwb.save('7. CHECK.xlsx')

wb = openpyxl.load_workbook('5. Shipping Report INPUT.xlsx')
fwb = openpyxl.load_workbook('7. Check.xlsx')
final_sheet = fwb['ShipRep_VIP']
sheet = wb['ShipRep_VIP']
delivery_column = 26
for row in range(2, sheet.max_row + 1):
    original_row = row
    delivery_cell = sheet.cell(row, delivery_column)
    overwrite_cell = final_sheet.cell(original_row, delivery_column)
    overwrite_cell.value = delivery_cell.value

fwb.save('6. Shipping Report Packlist OUTPUT.xlsx')