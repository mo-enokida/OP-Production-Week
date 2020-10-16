import openpyxl as xl
from openpyxl.styles import Font, Color, Fill, NamedStyle, PatternFill, Border, Side, Alignment, Protection
from openpyxl.styles import Font, NamedStyle, PatternFill, Border, Side, Alignment, Protection
from openpyxl.cell import Cell
from copy import copy, deepcopy
wb = xl.load_workbook('7. Weekly Ordering Index.xlsx')


#ADDS WEEKLY MEALS IN TO TEMPLATE SHEET

sheet = wb['1. MENU INDEX']
week_sheet = wb['THIS WEEK']
red_meal_input = input('RED POT: ')
teal_meal_input = input('TEAL POT: ')
green_meal_input = input('GREEN POT: ')
yellow_meal_input = input('YELLOW POT: ')
blue_meal_input = input('BLUE POT: ')


red_meal_input = int(red_meal_input)
teal_meal_input = int(teal_meal_input)
green_meal_input = int(green_meal_input)
yellow_meal_input = int(yellow_meal_input)
blue_meal_input = int(blue_meal_input)
meal_number_row = 1
week_paste_row = 1
red_column = 5
teal_column = 6
green_column = 8
yellow_column = 9
blue_column = 7
recipe_row = 60
recipe_column = 1


###IF FAMILY FAVORITES CHANGE, CHANGE THIS
fam_fave_recipes = ['#Butter Pickles','#Red Onion Marmalade','#Ketchup','#Remoulade','#Tomato Chili Sauce','#Taco Seasoning Mix','#Pickled Jalapenos','#Curtido Slaw','#Guacamole','#Pico de Gallo','#Pizza Dough','#Aged Red Wine Vinaigrette','#Marinara Sauce']
week_recipes = []
for row in range(1, 25):
    paste_cell = week_sheet.cell(row, red_column)
    copy_cell = sheet.cell(row + 1, red_meal_input)
    paste_cell.value = copy_cell.value
    ingredient = copy_cell.value
    if '#' in str(ingredient) and ingredient not in fam_fave_recipes:
        week_recipes.append(copy_cell.value)
for row in range(1, 25):
    paste_cell = week_sheet.cell(row, teal_column)
    copy_cell = sheet.cell(row + 1, teal_meal_input)
    paste_cell.value = copy_cell.value
    ingredient = copy_cell.value
    if '#' in str(ingredient) and ingredient not in fam_fave_recipes:
        week_recipes.append(copy_cell.value)
for row in range(1, 25):
    paste_cell = week_sheet.cell(row, green_column)
    copy_cell = sheet.cell(row + 1, green_meal_input)
    paste_cell.value = copy_cell.value
    ingredient = copy_cell.value
    if '#' in str(ingredient) and ingredient not in fam_fave_recipes:
        week_recipes.append(copy_cell.value)
for row in range(1, 25):
    paste_cell = week_sheet.cell(row, yellow_column)
    copy_cell = sheet.cell(row + 1, yellow_meal_input)
    paste_cell.value = copy_cell.value
    ingredient = copy_cell.value
    if '#' in str(ingredient) and ingredient not in fam_fave_recipes:
        week_recipes.append(copy_cell.value)
for row in range(1, 25):
    paste_cell = week_sheet.cell(row, blue_column)
    copy_cell = sheet.cell(row + 1, blue_meal_input)
    paste_cell.value = copy_cell.value
    ingredient = copy_cell.value
    if '#' in str(ingredient) and ingredient not in fam_fave_recipes:
        week_recipes.append(copy_cell.value)

recipes = list(dict.fromkeys(week_recipes))
print(recipes)


#ADDS EACH RECIPE INTO WEEK SHEET
recipe_index = wb['2. OP RECIPES INDEX']
paste_column = 1
for recipe in recipes:
    recipe_index_row = 2
    recipe_row = 60
    paste_column = paste_column + 1
    for column in range(1, recipe_index.max_column):
        recipe_cell = recipe_index.cell(recipe_index_row, column)
        recipe_cell = recipe_cell.value
        if recipe_cell == recipe:
            recipe_column = column
            for row in range(2, 40):
                copy_cell = recipe_index.cell(row, recipe_column)
                paste_cell = week_sheet.cell(row + 58, paste_column)
                paste_cell.value = copy_cell.value


ingredients = []
#ADDS INGREDIENTS FROM EACH MEAL
for column in range(1, week_sheet.max_column + 1):
    for row in range(2,29):
        Cell = week_sheet.cell(row, column)
        ingredient = Cell.value
        if '#' not in str(ingredient):
            ingredients.append(Cell.value)


#ADDS INGREDIENTS FROM FAM FAVORITE RECIPES

for column in range(1, week_sheet.max_column + 1):
    for row in range(30,59):
        Cell = week_sheet.cell(row, column)
        ingredient = Cell.value
        if '#' not in str(ingredient):
            ingredients.append(Cell.value)

#ADDS INGREDIENTS FROM WEEKLY RECIPES
for column in range(1, week_sheet.max_column + 1):
    for row in range(60,110):
        Cell = week_sheet.cell(row, column)
        ingredient = Cell.value
        if '#' not in str(ingredient):
            ingredients.append(Cell.value)


no_dupe_ingredients = list(dict.fromkeys(ingredients))
print(no_dupe_ingredients)

list_sheet = wb['AMT per RECIPE']

ingredient_column = 2
i = 1
row = 2
while i < len(no_dupe_ingredients):
    Cell = list_sheet.cell(row, ingredient_column)
    Cell.value = no_dupe_ingredients[i]
    i = i + 1
    row = row + 1


wb.save('8. Weekly Ordering Index.xlsx')

import openpyxl as xl
from openpyxl.styles import Font, Color, Fill, NamedStyle, PatternFill, Border, Side, Alignment, Protection
from openpyxl.styles import Font, NamedStyle, PatternFill, Border, Side, Alignment, Protection
from openpyxl.cell import Cell
from copy import copy, deepcopy
wb = xl.load_workbook('8. Weekly Ordering Index.xlsx')



####APPEND VENDOR LISTS
# ADD REST
sheet = wb['3. VENDORS INDEX']
cooseman = []
cooseman_column = 1
for row in range(2, sheet.max_row + 1):
    Cell = sheet.cell(row, cooseman_column)
    vendor_item = Cell.value
    cooseman.append(vendor_item)
if None in cooseman:
    cooseman.remove(None)

tama = []
tama_column = 2
for row in range(2, sheet.max_row + 1):
    Cell = sheet.cell(row, tama_column)
    vendor_item = Cell.value
    tama.append(vendor_item)
if None in tama:
    tama.remove(None)

sysco = []
sysco_column = 3
for row in range(2, sheet.max_row + 1):
    Cell = sheet.cell(row, sysco_column)
    vendor_item = Cell.value
    sysco.append(vendor_item)
if None in sysco:
    sysco.remove(None)

spices = []
spices_column = 4
for row in range(2, sheet.max_row + 1):
    Cell = sheet.cell(row, spices_column)
    vendor_item = Cell.value
    spices.append(vendor_item)
if None in spices:
    spices.remove(None)


torn_glasser = []
torn_column = 5
for row in range(2, sheet.max_row + 1):
    Cell = sheet.cell(row, torn_column)
    vendor_item = Cell.value
    torn_glasser.append(vendor_item)
if None in torn_glasser:
    torn_glasser.remove(None)

smart_final = []
smartfinal_column = 6
for row in range(2, sheet.max_row + 1):
    Cell = sheet.cell(row, smartfinal_column)
    vendor_item = Cell.value
    smart_final.append(vendor_item)
if None in smart_final:
    smart_final.remove(None)

guerrero = []
guerrero_column = 7
for row in range(2, sheet.max_row + 1):
    Cell = sheet.cell(row, guerrero_column)
    vendor_item = Cell.value
    guerrero.append(vendor_item)
if None in guerrero:
    guerrero.remove(None)

laxc = []
laxc_column = 8
for row in range(2, sheet.max_row + 1):
    Cell = sheet.cell(row, laxc_column)
    vendor_item = Cell.value
    laxc.append(vendor_item)
if None in laxc:
    laxc.remove(None)

mutual_trading = []
mt_column = 9
for row in range(2, sheet.max_row + 1):
    Cell = sheet.cell(row, mt_column)
    vendor_item = Cell.value
    mutual_trading.append(vendor_item)
if None in mutual_trading:
    mutual_trading.remove(None)

chefs_warehouse = []
chefswarehouse_column = 10
for row in range(2, sheet.max_row + 1):
    Cell = sheet.cell(row, chefswarehouse_column)
    vendor_item = Cell.value
    chefs_warehouse.append(vendor_item)
if None in chefs_warehouse:
    chefs_warehouse.remove(None)

restaurant_depot = []
misc_column = 11
for row in range(2, sheet.max_row + 1):
    Cell = sheet.cell(row, misc_column)
    vendor_item = Cell.value
    restaurant_depot.append(vendor_item)
if None in restaurant_depot:
    restaurant_depot.remove(None)

pasta_mia = []
misc_column = 12
for row in range(2, sheet.max_row + 1):
    Cell = sheet.cell(row, misc_column)
    vendor_item = Cell.value
    pasta_mia.append(vendor_item)
if None in pasta_mia:
    pasta_mia.remove(None)


sheet = wb['AMT per RECIPE']
ingredient_column = 2
vendor_column = 1
for row in range(2, sheet.max_row + 1):
    ingredient = sheet.cell(row, ingredient_column)
    vendor = sheet.cell(row, vendor_column)
    vendor.value = ''
    if ingredient.value in cooseman:
        vendor.value = 'cooseman,'
    if ingredient.value in tama:
        vendor.value = vendor.value + 'tama,'
    if ingredient.value in sysco:
        vendor.value = vendor.value + 'sysco,'
    if ingredient.value in spices:
        vendor.value = vendor.value + 'spices,'
    if ingredient.value in torn_glasser:
        vendor.value = vendor.value + 'torn_glasser,'
    if ingredient.value in smart_final:
        vendor.value = vendor.value + 'smart_final,'
    if ingredient.value in guerrero:
        vendor.value = vendor.value + 'guerrero,'
    if ingredient.value in laxc:
        vendor.value = vendor.value + 'laxc,'
    if ingredient.value in mutual_trading:
        vendor.value = vendor.value + 'mutual_trading,'
    if ingredient.value in chefs_warehouse:
        vendor.value = vendor.value + 'chefs_warehouse,'
    if ingredient.value in restaurant_depot:
        vendor.value = vendor.value + 'restaurant_depot,'
    if ingredient.value in pasta_mia:
        vendor.value = vendor.value + 'pasta_mia,'

sheet = wb['AMT per RECIPE']
ingredient_column = 2
vendor_column = 1
for row in range(2, sheet.max_row + 1):
    ingredient = sheet.cell(row, ingredient_column)
    vendor = sheet.cell(row, vendor_column)
    if vendor.value == '':
        vendor.value = 'OTHER'
wb.save('8. Weekly Ordering Index.xlsx')


